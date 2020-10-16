import openpyxl as xl
import sys
from collections import namedtuple

try:
    wb = xl.load_workbook("ark.xlsx")
    sheet = wb.active # Use Default sheet
except:
    print("Failed to open workbook")
    sys.exit(1)

# Set some vars
ROWS = sheet.max_row
QUESTION_COL = 3
ANSWER_COL = 3
CORRECT_COL = 11
ANSWERS_PER_QUESTION = 4

# This will grab the questions
Answer = namedtuple("Answer", ["ID", "Text", "Correct"])
class Question():
    def __init__(self, questionID, questionText):
        self.questionID = questionID
        self.questionText = questionText
        self.answers = []
    def add_answer(self, answer):
        """Please give answer as a NamedTuple object"""
        self.answers.append(answer)
    def __str__(self):
        return "{:02d}: {}".format(self.questionID, self.questionText)

logged_questions = []

# Loop over the sheet
for i in range(1, ROWS):
    # Check if we have a new question
    cell = sheet.cell(row=i, column=1)
    if cell.value is not None and cell.value.startswith("Spørgsmål"):
        # We need to record a new question
        qText = sheet.cell(row=i, column=QUESTION_COL).value
        if qText is None:
            print("Error on row {} no question text".format(i))
        # Extract id
        qId = int(cell.value.split(" ")[-1])
        question = Question(qId, qText)

        # Now to extract the answers
        j = i + 1
        readingAnswers = False
        finishedReading = False
        while not finishedReading:
            # Read a cell
            c = sheet.cell(row=j, column=ANSWER_COL)
            # Control behaviour
            if readingAnswers == False and c.value is not None:
                readingAnswers = True
            if readingAnswers == True and c.value is None:
                finishedReading = True
                break
            # Read an answer into a tuple
            if readingAnswers:
                aText = c.value
                aId = int(sheet.cell(row=j, column=ANSWER_COL - 1).value[:-1])
                aCorrect = sheet.cell(row=j, column=CORRECT_COL).value is not None
                answer = Answer(aId, aText, aCorrect)
                # Add to our question
                question.add_answer(answer)
            j += 1
        logged_questions.append(question)

# Asser do something here